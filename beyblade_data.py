#!/usr/bin/env python3
import hashlib
import io
import json
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from pathlib import Path
from datetime import datetime, timezone
import openpyxl
from PIL import Image as PILImage

# --- Constants ---
PREFIX_TO_SHEET = {
    "blades": "Parts_X_Blades", "ratchets": "Parts_X_Ratchets", "bits": "Parts_X_Bits",
    "lockChips": "Parts_X_LockChips", "overBlades": "Parts_X_OverBlades",
    "metalBlades": "Parts_X_MetalBlades", "mainBlades": "Parts_X_MainBlades",
    "assistBlades": "Parts_X_AssistBlades",
}
SHEET_SETTINGS = {
    "Products": ("productID", "boxImageUrl", "minifiedBoxImage", "products"),
    "Accessories": ("dataID", "imageUrl", "minifiedImage", "accessories"),
}
NO_IMAGE_SHEETS = {"ProductContents": "productContents"}

WEBP_QUALITY = 85

# --- Core Logic Helpers ---

def read_sheet_rows(ws):
    """Stops at the first fully-blank row."""
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            break
        rows.append(dict(zip(headers, row)))
    return rows

def png_bytes_to_webp(png_bytes: bytes) -> bytes:
    img = PILImage.open(io.BytesIO(png_bytes)).convert("RGBA")
    buf = io.BytesIO()
    img.save(buf, format="WEBP", quality=WEBP_QUALITY)
    return buf.getvalue()

def extract_embedded_images(ws, id_col_1idx: int, img_col_1idx: int, min_col_1idx: int):
    """Returns {row_number_1idx: {'full': png_bytes, 'min': png_bytes}}"""
    result = {}
    for img in ws._images:
        row = img.anchor._from.row + 1
        col = img.anchor._from.col + 1
        raw = img._data()
        if col == img_col_1idx:
            result.setdefault(row, {})["full"] = raw
        elif col == min_col_1idx:
            result.setdefault(row, {})["min"] = raw
    return result

def build_jsdelivr_url(username: str, repo: str, branch: str = "main") -> str:
    """Constructed from three plain, separately-validated pieces --
    never hand-typed as one long URL, which is exactly how a stray
    character (a pasted newline, a typo) ends up baked into every single
    image URL in the output JSON without anyone noticing until images
    fail to load."""
    return f"https://cdn.jsdelivr.net/gh/{username}/{repo}@{branch}/data/"

# --- UI Application ---

class PackagerEmbeddedApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Offline Packager (From Embedded Images)")
        self.root.geometry("750x650")
        self.root.minsize(650, 500)

        # Variables
        self.excel_path_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.github_username_var = tk.StringVar()
        self.github_repo_var = tk.StringVar()
        self.github_branch_var = tk.StringVar(value="main")

        self._build_ui()

    def _build_ui(self):
        # --- Config Frame ---
        config_frame = ttk.LabelFrame(self.root, text="Configuration", padding=(10, 10))
        config_frame.pack(fill=tk.X, padx=10, pady=10)

        # 1. Excel File
        ttk.Label(config_frame, text="Master Excel File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.excel_path_var, width=50).grid(row=0, column=1, sticky=tk.EW, padx=5)
        ttk.Button(config_frame, text="Browse...", command=self._browse_excel).grid(row=0, column=2, padx=5)

        # 2. Output Folder
        ttk.Label(config_frame, text="Output Folder:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.output_dir_var, width=50).grid(row=1, column=1, sticky=tk.EW, padx=5)
        ttk.Button(config_frame, text="Browse...", command=self._browse_output).grid(row=1, column=2, padx=5)

        # 3. GitHub Username / Repo / Branch -- three small, plain fields
        # instead of one long URL field. The jsDelivr URL gets built
        # programmatically from these; nothing here is ever hand-typed
        # as a single complex string.
        ttk.Label(config_frame, text="GitHub Username:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.github_username_var, width=30).grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(config_frame, text="Repo Name:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.github_repo_var, width=30).grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

        ttk.Label(config_frame, text="Branch:").grid(row=4, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.github_branch_var, width=15).grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)

        self.preview_label = ttk.Label(config_frame, text="", font=("", 8), foreground="#0066cc", wraplength=600)
        self.preview_label.grid(row=5, column=0, columnspan=3, sticky=tk.W, padx=5, pady=(0, 5))
        for var in (self.github_username_var, self.github_repo_var, self.github_branch_var):
            var.trace_add("write", lambda *_: self._update_preview())
        self._update_preview()

        config_frame.columnconfigure(1, weight=1)

        # --- Action Frame ---
        action_frame = ttk.Frame(self.root)
        action_frame.pack(fill=tk.X, padx=10, pady=5)

        self.run_btn = ttk.Button(action_frame, text="Run Packager", command=self._start_processing)
        self.run_btn.pack(side=tk.LEFT)

        self.progress = ttk.Progressbar(action_frame, mode='indeterminate')

        # --- Log Frame ---
        log_frame = ttk.LabelFrame(self.root, text="Logs", padding=(10, 10))
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.log_text = ScrolledText(log_frame, state='disabled', bg="#f4f4f4", wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _update_preview(self):
        username = self.github_username_var.get().strip()
        repo = self.github_repo_var.get().strip()
        branch = self.github_branch_var.get().strip() or "main"
        if username and repo:
            self.preview_label.config(text=f"Resulting URL prefix: {build_jsdelivr_url(username, repo, branch)}")
        else:
            self.preview_label.config(text="Enter username and repo to preview the resulting URL.")

    def _browse_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if path: self.excel_path_var.set(path)

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select Output Data Folder")
        if path: self.output_dir_var.set(path)

    def log(self, message):
        """Thread-safe UI logging."""
        def append():
            self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state='disabled')
        self.root.after(0, append)

    def _start_processing(self):
        # Raw values, checked BEFORE any stripping -- this is the actual
        # fix. Checking post-strip values (what was here before) can
        # never catch anything: a trailing newline is exactly what
        # .strip() removes, so by the time you compare a stripped value
        # against itself, the evidence is already gone.
        raw_username = self.github_username_var.get()
        raw_repo = self.github_repo_var.get()
        raw_branch = self.github_branch_var.get()

        for label, raw_value in (("username", raw_username), ("repo", raw_repo), ("branch", raw_branch)):
            if raw_value and (raw_value != raw_value.strip() or "\n" in raw_value or "\r" in raw_value):
                messagebox.showerror(
                    "Suspicious Input",
                    f"The {label} field has a newline or leading/trailing whitespace in it -- this would "
                    f"corrupt every image URL in the output. Please retype it directly instead of pasting."
                )
                return

        username = raw_username.strip()
        repo = raw_repo.strip()
        branch = raw_branch.strip() or "main"

        if not all([self.excel_path_var.get(), self.output_dir_var.get(), username, repo]):
            messagebox.showerror("Missing Inputs", "Please provide the Excel file, output folder, GitHub username, and repo name.")
            return

        base_url = build_jsdelivr_url(username, repo, branch)

        # Prepare UI
        self.run_btn.config(state=tk.DISABLED)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        self.progress.start(15)
        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')

        self.log(f"Using base URL: {base_url}")

        # Run script in background thread
        threading.Thread(target=self._run_script, args=(base_url,), daemon=True).start()

    def _run_script(self, base_url: str):
        try:
            excel_path = Path(self.excel_path_var.get())
            out_dir = Path(self.output_dir_var.get())

            self._package(excel_path, out_dir, base_url, self.log)

            self.log("\nSUCCESS: Packaging completed successfully.")
            self.root.after(0, lambda: messagebox.showinfo("Done", "Packaging completed successfully!"))

        except Exception as e:
            import traceback
            err = traceback.format_exc()
            self.log(f"\nERROR:\n{err}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred:\n{e}"))
        finally:
            self.root.after(0, self._stop_processing)

    def _stop_processing(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.run_btn.config(state=tk.NORMAL)

    def _process_sheet(self, wb, sheet_name, prefix, id_field, img_field, min_field, out_dir, raw_base_url, images_by_generation, manifest):
        if sheet_name not in wb.sheetnames:
            self.log(f"  SKIP (not in workbook): {sheet_name}")
            return

        ws = wb[sheet_name]
        headers = [str(c.value) if c.value else "" for c in ws[1]]

        if id_field not in headers or img_field not in headers or min_field not in headers:
            self.log(f"  SKIP {sheet_name}: missing expected columns")
            return

        id_col = headers.index(id_field) + 1
        img_col = headers.index(img_field) + 1
        min_col = headers.index(min_field) + 1

        rows = read_sheet_rows(ws)
        images_by_row = extract_embedded_images(ws, id_col, img_col, min_col)
        out_images_dir = out_dir / "images"
        out_images_dir.mkdir(parents=True, exist_ok=True)

        with_image = without_image = 0
        for i, row in enumerate(rows):
            data_id = row.get(id_field)
            if data_id is None:
                continue

            sheet_row_1idx = i + 2  # +2: header is row 1, rows list is 0-indexed from row 2
            images = images_by_row.get(sheet_row_1idx, {})

            row.pop(img_field, None)
            row.pop(min_field, None)

            if "full" in images:
                webp_bytes = png_bytes_to_webp(images["full"])
                full_name = f"{prefix}-{data_id}.webp"
                (out_images_dir / full_name).write_bytes(webp_bytes)
                row["imageUrl"] = f"{raw_base_url}images/{full_name}"
                row["imageHash"] = hashlib.sha1(webp_bytes).hexdigest()[:12]
                with_image += 1
                generation = row.get("hobbyGeneration", "unknown")
                images_by_generation.setdefault(generation, {}).setdefault(prefix, []).append(str(data_id))
            else:
                row["imageUrl"] = None
                row["imageHash"] = None
                without_image += 1

            if "min" in images:
                webp_bytes = png_bytes_to_webp(images["min"])
                min_name = f"{prefix}-{data_id}-min.webp"
                (out_images_dir / min_name).write_bytes(webp_bytes)
                row["minifiedImageUrl"] = f"{raw_base_url}images/{min_name}"
                row["minifiedImageHash"] = hashlib.sha1(webp_bytes).hexdigest()[:12]
            else:
                row["minifiedImageUrl"] = None
                row["minifiedImageHash"] = None

        out_path = out_dir / f"{prefix}.json"
        out_path.write_text(json.dumps(rows, separators=(",", ":"), ensure_ascii=False), encoding="utf-8")
        manifest["sheets"][prefix] = {"rowCount": len(rows), "withImage": with_image, "withoutImage": without_image}
        self.log(f"  {sheet_name} -> {out_path.name}: {len(rows)} rows ({with_image} with image, {without_image} without)")

    def _package(self, xlsx_path: Path, out_dir: Path, raw_base_url: str, log):
        """The core package logic adapted from the script."""
        if not raw_base_url.endswith("/"):
            raw_base_url += "/"
        out_dir.mkdir(parents=True, exist_ok=True)

        log(f"Loading workbook (this may take a moment): {xlsx_path.name}...")

        # data_only=True is required so formula cells are resolved to their values.
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        manifest = {"packagedAt": datetime.now(timezone.utc).isoformat(), "sheets": {}}
        images_by_generation = {}

        for prefix, sheet_name in PREFIX_TO_SHEET.items():
            self._process_sheet(wb, sheet_name, prefix, "dataID", "image", "minifiedImage", out_dir, raw_base_url, images_by_generation, manifest)

        for sheet_name, (id_field, img_field, min_field, prefix) in SHEET_SETTINGS.items():
            self._process_sheet(wb, sheet_name, prefix, id_field, img_field, min_field, out_dir, raw_base_url, images_by_generation, manifest)

        for sheet_name, short_key in NO_IMAGE_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            rows = read_sheet_rows(wb[sheet_name])
            (out_dir / f"{short_key}.json").write_text(json.dumps(rows, separators=(",", ":"), ensure_ascii=False), encoding="utf-8")
            manifest["sheets"][short_key] = {"rowCount": len(rows)}
            log(f"  {sheet_name} -> {short_key}.json: {len(rows)} rows (no images)")

        manifest["imagesByGeneration"] = {g: {k: len(v) for k, v in s.items()} for g, s in images_by_generation.items()}
        (out_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        (out_dir / "_images_by_generation.json").write_text(json.dumps(images_by_generation), encoding="utf-8")
        log(f"\nDone. Output in {out_dir}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PackagerEmbeddedApp(root)
    root.mainloop()